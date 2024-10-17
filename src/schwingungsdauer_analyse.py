import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook

# Dateinamen definieren
script_name = "schwingungsdauer_analyse.py"
excel_name = "schwingungsdauer_messungen.xlsx"

print(f"Dieses Skript heißt {script_name}")

# Generieren der Schwingungsdauer-Daten
np.random.seed(42)  # Für Reproduzierbarkeit
n_measurements = 10
true_period = 2.00  # Wahre Schwingungsdauer in Sekunden
noise = 0.05  # Standardabweichung des Rauschens

measurements = np.random.normal(true_period, noise, n_measurements)
measurements = np.round(measurements, 2)  # Runden auf 2 Dezimalstellen

# Erstellen und Speichern der Excel-Datei
wb = Workbook()
ws = wb.active
ws.title = "Schwingungsdauer"

ws['A1'] = "Messung"
ws['B1'] = "T in s"

for i, measurement in enumerate(measurements, start=1):
    ws[f'A{i+1}'] = i
    ws[f'B{i+1}'] = measurement

wb.save(excel_name)
print(f"Datei {excel_name} wurde erstellt.")

# Einlesen der Excel-Datei
df = pd.read_excel(excel_name)

# Extrahieren der Messwerte aus der zweiten Spalte
T = df.iloc[:, 1].values

# Berechnung des Mittelwerts
T_mean = np.mean(T)

# Berechnung der Standardabweichung
T_std = np.std(T, ddof=1)  # ddof=1 für Stichproben-Standardabweichung

print(f"Mittlere Schwingungsdauer: {T_mean:.2f} s")
print(f"Standardabweichung: {T_std:.2f} s")

# Erstellen eines Histogramms
plt.figure(figsize=(10, 6))
plt.hist(T, bins='auto', edgecolor='black')
plt.title('Verteilung der Schwingungsdauer-Messungen')
plt.xlabel('Schwingungsdauer (s)')
plt.ylabel('Häufigkeit')
plt.grid(True, alpha=0.3)
plt.show()

# Erstellen eines Streudiagramms
plt.figure(figsize=(10, 6))
plt.scatter(range(1, n_measurements + 1), T, color='blue')
plt.axhline(y=T_mean, color='red', linestyle='--', label=f'Mittelwert: {T_mean:.2f} s')
plt.fill_between(range(1, n_measurements + 1), T_mean - T_std, T_mean + T_std, 
                 color='gray', alpha=0.2, label=f'Standardabweichung: ±{T_std:.2f} s')
plt.title('Streudiagramm der Schwingungsdauer-Messungen')
plt.xlabel('Messung')
plt.ylabel('Schwingungsdauer (s)')
plt.legend()
plt.grid(True, alpha=0.3)
plt.show()